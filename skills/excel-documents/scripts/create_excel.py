#!/usr/bin/env python3
"""
create_excel.py – Erstellt professionelle Excel-Dateien (.xlsx) mit raspb-Styling.

Usage:
  python3 create_excel.py --json /tmp/excel_data.json --output /tmp/output.xlsx
  python3 create_excel.py --json /tmp/excel_data.json --output /tmp/output.xlsx --template /path/to/template.xlsx

JSON-Format:
{
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["Spalte A", "Spalte B", "Spalte C"],
      "rows": [
        ["Wert 1", 100, "2026-03-20"],
        ["Wert 2", 200, "2026-03-21"]
      ],
      "column_widths": [30, 15, 20],
      "number_formats": {"B": "#,##0.00 €", "C": "DD.MM.YYYY"},
      "formulas": {"B": "SUM"},
      "freeze_panes": "A2",
      "autofilter": true
    }
  ],
  "title": "Mein Dokument",
  "author": "raspb Webservices UG"
}
"""
import json
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

# raspb Design System Colors
PINK = "E8458B"
ICONIC_BLUE = "121D33"
DARK_GREY = "333333"
LIGHT_GREY = "EEEEEE"
BORDER_GREY = "DDDDDD"
WHITE = "FFFFFF"

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", color=DARK_GREY, size=11)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

ALT_ROW_FILL = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_GREY),
    right=Side(style="thin", color=BORDER_GREY),
    top=Side(style="thin", color=BORDER_GREY),
    bottom=Side(style="thin", color=BORDER_GREY),
)

SUMMARY_FONT = Font(name="Calibri", bold=True, color=ICONIC_BLUE, size=11)
SUMMARY_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")


def apply_cell_style(cell, font, fill=None, alignment=None, border=None, number_format=None):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def create_sheet(wb, sheet_config, index=0):
    if index == 0:
        ws = wb.active
        ws.title = sheet_config.get("name", "Sheet1")
    else:
        ws = wb.create_sheet(title=sheet_config.get("name", f"Sheet{index + 1}"))

    headers = sheet_config.get("headers", [])
    rows = sheet_config.get("rows", [])
    col_widths = sheet_config.get("column_widths", [])
    num_formats = sheet_config.get("number_formats", {})
    formulas = sheet_config.get("formulas", {})
    freeze = sheet_config.get("freeze_panes", None)
    autofilter = sheet_config.get("autofilter", False)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL, HEADER_ALIGNMENT, THIN_BORDER)

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            col_letter = get_column_letter(col_idx)

            nf = num_formats.get(col_letter, None)
            apply_cell_style(cell, DATA_FONT, fill, DATA_ALIGNMENT, THIN_BORDER, nf)

    # Summary row with formulas
    if formulas and rows:
        summary_row = len(rows) + 2
        for col_letter, formula_type in formulas.items():
            col_idx = ord(col_letter.upper()) - ord("A") + 1
            data_start = 2
            data_end = len(rows) + 1

            if formula_type.upper() == "SUM":
                formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            elif formula_type.upper() == "AVG" or formula_type.upper() == "AVERAGE":
                formula = f"=AVERAGE({col_letter}{data_start}:{col_letter}{data_end})"
            elif formula_type.upper() == "COUNT":
                formula = f"=COUNTA({col_letter}{data_start}:{col_letter}{data_end})"
            elif formula_type.upper() == "MIN":
                formula = f"=MIN({col_letter}{data_start}:{col_letter}{data_end})"
            elif formula_type.upper() == "MAX":
                formula = f"=MAX({col_letter}{data_start}:{col_letter}{data_end})"
            else:
                formula = formula_type  # Custom formula

            cell = ws.cell(row=summary_row, column=col_idx, value=formula)
            nf = num_formats.get(col_letter, None)
            apply_cell_style(cell, SUMMARY_FONT, SUMMARY_FILL, DATA_ALIGNMENT, THIN_BORDER, nf)

        # Label for summary row
        label_cell = ws.cell(row=summary_row, column=1, value="Gesamt")
        apply_cell_style(label_cell, SUMMARY_FONT, SUMMARY_FILL, DATA_ALIGNMENT, THIN_BORDER)

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-width for columns without explicit width
    if not col_widths:
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8
            for row_idx in range(2, len(rows) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze panes
    if freeze:
        ws.freeze_panes = freeze

    # Autofilter
    if autofilter and headers:
        last_col = get_column_letter(len(headers))
        last_row = len(rows) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return ws


def main():
    parser = argparse.ArgumentParser(description="Create professional Excel files with raspb styling")
    parser.add_argument("--json", required=True, help="Path to JSON config file")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--template", help="Optional template .xlsx to use as base")
    args = parser.parse_args()

    # Load JSON config
    with open(args.json, "r", encoding="utf-8") as f:
        config = json.load(f)

    # Create workbook
    if args.template:
        wb = load_workbook(args.template)
    else:
        wb = Workbook()

    # Set document properties
    wb.properties.title = config.get("title", "raspb Document")
    wb.properties.creator = config.get("author", "raspb Webservices UG")
    wb.properties.company = "raspb Webservices UG"

    # Create sheets
    sheets = config.get("sheets", [])
    if not sheets:
        print("Error: No sheets defined in JSON config")
        sys.exit(1)

    for idx, sheet_config in enumerate(sheets):
        create_sheet(wb, sheet_config, idx)

    # Remove default empty sheet if we loaded a template
    if args.template and "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            del wb["Sheet"]
        except KeyError:
            pass

    # Save
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Success: {output_path}")


if __name__ == "__main__":
    main()
